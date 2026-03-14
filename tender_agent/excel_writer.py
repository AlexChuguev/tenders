from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from tender_agent.models import TenderAnalysisResult


SUMMARY_TAGS = [
    "Не наш тип работ",
    "Частично наш стек",
    "Не наш стек",
    "Не подходим по требованиям",
    "Мало опыта / кейсов",
    "Предварительно подходит",
    "В работу",
    "Подано",
    "Не успели обработать",
    "Не изучено",
]

DETAIL_HEADERS = [
    "Статус",
    "Уверенность, %",
    "Комментарий",
    "№",
    "Наименование лота",
    "Дата окончания приема заявок",
]

MONTH_NAMES = {
    1: "Январь",
    2: "Февраль",
    3: "Март",
    4: "Апрель",
    5: "Май",
    6: "Июнь",
    7: "Июль",
    8: "Август",
    9: "Сентябрь",
    10: "Октябрь",
    11: "Ноябрь",
    12: "Декабрь",
}


class ExcelWriter:
    def __init__(self, output_path: Path) -> None:
        self.output_path = output_path
        self.output_path.parent.mkdir(parents=True, exist_ok=True)

    def ensure_header(self) -> None:
        workbook = self._open()
        self._cleanup_legacy_sheet(workbook)
        workbook.save(self.output_path)

    def append_result(self, result: TenderAnalysisResult) -> None:
        workbook = self._open()
        self._cleanup_legacy_sheet(workbook)
        worksheet = self._ensure_month_sheet(workbook, result.deadline_at)
        self._remove_existing_row(worksheet, result.tender_id)

        row_number = max(worksheet.max_row + 1, len(SUMMARY_TAGS) + 2)
        detail_number = row_number - len(SUMMARY_TAGS) - 1
        worksheet.append(
            [
                "",
                "",
                self._normalize_tag_for_sheet(result),
                result.confidence_percent,
                result.classification_comment,
                detail_number,
                result.title,
                _format_deadline(result.deadline_at),
            ]
        )
        title_cell = worksheet.cell(row=worksheet.max_row, column=7)
        title_cell.hyperlink = result.url
        title_cell.style = "Hyperlink"
        workbook.save(self.output_path)

    def _open(self):
        if self.output_path.exists():
            return load_workbook(self.output_path)
        workbook = Workbook()
        workbook.active.title = "Temp"
        return workbook

    def _cleanup_legacy_sheet(self, workbook) -> None:
        if "Tenders" in workbook.sheetnames and len(workbook.sheetnames) > 1:
            del workbook["Tenders"]
        if "Temp" in workbook.sheetnames and len(workbook.sheetnames) > 1:
            del workbook["Temp"]

    def _ensure_month_sheet(self, workbook, deadline_at: datetime | None):
        sheet_name = _sheet_name(deadline_at)
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            self._ensure_sheet_structure(worksheet)
            return worksheet

        worksheet = workbook.create_sheet(title=sheet_name)
        self._ensure_sheet_structure(worksheet)
        return worksheet

    def _ensure_sheet_structure(self, worksheet) -> None:
        worksheet.cell(row=1, column=1, value="Общие итоги")
        worksheet.cell(row=1, column=2, value="Кол-во")
        for index, header in enumerate(DETAIL_HEADERS, start=3):
            worksheet.cell(row=1, column=index, value=header)
        for row_index, tag in enumerate(SUMMARY_TAGS, start=2):
            worksheet.cell(row=row_index, column=1, value=tag)
            worksheet.cell(row=row_index, column=2, value=f'=COUNTIF(C:C,A{row_index})')
        self._set_widths(worksheet)

    def _set_widths(self, worksheet) -> None:
        widths = {
            "A": 28,
            "B": 12,
            "C": 28,
            "D": 14,
            "E": 48,
            "F": 10,
            "G": 110,
            "H": 18,
        }
        for column, width in widths.items():
            worksheet.column_dimensions[column].width = width

    def _normalize_tag_for_sheet(self, result: TenderAnalysisResult) -> str:
        return result.classification_tag or "Не изучено"

    def _remove_existing_row(self, worksheet, tender_id: str) -> None:
        for row_index in range(len(SUMMARY_TAGS) + 2, worksheet.max_row + 1):
            hyperlink = worksheet.cell(row=row_index, column=7).hyperlink
            target = hyperlink.target if hyperlink else ""
            if isinstance(target, str) and tender_id and tender_id in target:
                worksheet.delete_rows(row_index, 1)
                break


def _sheet_name(deadline_at: datetime | None) -> str:
    if deadline_at is None:
        return "Без даты"
    return f"{MONTH_NAMES[deadline_at.month]} {deadline_at.year}"


def _format_deadline(deadline_at: datetime | None) -> str:
    if deadline_at is None:
        return ""
    if deadline_at.hour == 0 and deadline_at.minute == 0:
        return deadline_at.strftime("%-d.%-m.%Y")
    return deadline_at.strftime("%-d.%-m.%Y %H:%M")
